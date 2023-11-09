# 10-1:
import openpyxl

wb = openpyxl.Workbook()  # Создаю новую таблицу
ws1 = wb.create_sheet('Лист1')

ws1['A1'].value = 'ФИО'
ws1['B1'].value = 'Результат'
ws1['A2'].value = 'Иванов'
ws1['A3'].value = 'Петров'
ws1['A4'].value = 'Иванов'
ws1['B2'].value = '100'
ws1['B3'].value = '400'
ws1['B4'].value = '200'

wb.create_sheet('Лист2')
ws2 = wb['Лист2']  # Мы не переключаемся, а указываем что в этой переменной лежит лист, куда нужно вносит информацию, когда мы обращаемся к этой переменной

result = {}

# iter_rows встроенная функция, которая идет по строчкам
for row in ws1.iter_rows(values_only=True):
    name = row[0]
    work = row[1]
    if name != 'ФИО':  # Пропускаем шапку
        if name in result:
            result[name] += int(work)
        else:
            result[name] = int(work)
# Если в словаре уже есть имя, то добавь значение, если имени нет, то добавь и имя и значение
ws2['A1'].value = 'ФИО'
ws2['B1'].value = 'Результат'
row_num = 2  # Шапку пропускаем

# Одновременно идет и по словарю и по коонкам во втором листе
for name, work in result.items():
    ws2.cell(row=row_num, column=1, value=name)
    ws2.cell(row=row_num, column=2, value=work)
    row_num += 1

ws2.cell(row=row_num, column=1, value='ИТОГО:')
ws2.cell(row=row_num, column=2, value=sum(result.values()))
# Value = сумма всех значений словаря

wb.save('Task10_1.xlsx')


# 10-2:
wb = openpyxl.Workbook()

ws1 = wb.create_sheet('Лист1')
ws1['A1'].value = 'Фамилия'
ws1['B1'].value = 'Результат'
ws1['A2'].value = 'Иванов'
ws1['A3'].value = 'Петров'
ws1['A4'].value = 'Сидоров'
ws1['A5'].value = 'Жуков'
ws1['B2'].value = '100'
ws1['B3'].value = '400'
ws1['B4'].value = '200'
ws1['B5'].value = '200'

wb.create_sheet('Лист2')
ws2 = wb['Лист2']
ws2['A1'].value = 'Фамилия'
ws2['B1'].value = 'Результат'
ws2['A2'].value = 'Андреев'
ws2['A3'].value = 'Петров'
ws2['A4'].value = 'Сидоров'
ws2['A5'].value = 'Жуков'
ws2['B2'].value = '150'
ws2['B3'].value = '400'
ws2['B4'].value = '200'
ws2['B5'].value = '200'

wb.create_sheet('Лист3')
ws3 = wb['Лист3']
ws3['A1'].value = 'Фамилия'
ws3['B1'].value = 'Результат'

result = {}

for row in ws1.iter_rows(values_only=True):
    name = row[0]
    work = row[1]
    if name != 'Фамилия':
        if name in result:
            result[name] += int(work)
        else:
            result[name] = int(work)

for row in ws2.iter_rows(values_only=True):
    name = row[0]
    work = row[1]
    if name != 'Фамилия':
        if name in result:
            result[name] += int(work)
        else:
            result[name] = int(work)

# Если мы применяем сортед, то надо еще раз применить dict, чтобы в итоге получился опять словарь

sorted_result = dict(sorted(result.items()))

num_row = 2
for name, work in sorted_result.items():
    ws3.cell(row=num_row, column=1, value=name)
    ws3.cell(row=num_row, column=2, value=work)
    num_row += 1

wb.save('Task10_2.xlsx')


# 10-3:
wb = openpyxl.Workbook()
ws1 = wb.active

ws1['A1'].value = 'Фамилия'
ws1['B1'].value = 'Результат'

ws1['A2'].value = 'Иванов'
ws1['A3'].value = 'Петров'
ws1['A4'].value = 'Сидоров'
ws1['A5'].value = 'Жуков'
ws1['A6'].value = 'Андреев'

ws1['B2'].value = '100'
ws1['B3'].value = '400'
ws1['B4'].value = '200'
ws1['B5'].value = '200'
ws1['B5'].value = '75'
ws1['B6'].value = '115'

wb.create_sheet('Лист2')
ws2 = wb['Лист2']

ws2['A1'].value = 'Минимальное'
ws2['A2'].value = 'Максимальное'
ws2['A3'].value = 'Сред. арефм.'
ws2['A4'].value = 'Медиана'

result = []
for row in ws1.iter_rows(values_only=True):
    name = row[0]
    work = row[1]
    if name != 'Фамилия':
        result.append(int(work))

ws2['B1'].value = min(result)
ws2['B2'].value = max(result)
ws2['B3'].value = sum(result)/len(result)
result.sort()

if len(result) % 2 != 0:
    med = result[len(result)//2]
    ws2['B4'].value = med
else:
    med = (result[len(result) // 2 - 1] + result[len(result) // 2]) / 2
    ws2['B4'].value = med

wb.save('Task10_3.xlsx')
